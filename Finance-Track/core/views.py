from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.template.loader import render_to_string
from decimal import Decimal
import json
from .models import Team, Account, Transaction, Category, Currency
from .forms import TransactionForm, AccountForm, CategoryForm, TeamForm


@login_required
def dashboard(request):
    """Main dashboard view showing overview of finances"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q, Count
    
    # Handle USD rate update
    if request.method == 'POST' and request.POST.get('update_usd_rate'):
        try:
            new_rate = Decimal(request.POST.get('usd_exchange_rate', '0'))
            if new_rate > 0:
                usd_currency = Currency.objects.get(code='USD')
                old_rate = usd_currency.exchange_rate_to_pkr
                usd_currency.exchange_rate_to_pkr = new_rate
                usd_currency.save()
                
                # Recalculate balances for all USD accounts
                usd_accounts = Account.objects.filter(currency=usd_currency)
                for account in usd_accounts:
                    account.calculate_current_balance()
                    account.save()
                
                # Update USD transaction PKR amounts
                usd_transactions = Transaction.objects.filter(currency=usd_currency)
                for transaction in usd_transactions:
                    transaction.amount_pkr = transaction.amount * new_rate
                    transaction.save(update_fields=['amount_pkr'])
                
                if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'USD rate updated from {old_rate} to {new_rate} PKR. All balances recalculated!'
                    })
            else:
                if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Invalid exchange rate'})
        except (Currency.DoesNotExist, ValueError, TypeError):
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Error updating exchange rate'})
    
    user_teams = request.user.teams.all()
    
    # Get total balances in PKR across all accounts
    total_balance_pkr = Account.objects.filter(
        is_active=True
    ).aggregate(total=Sum('current_balance_pkr'))['total'] or 0
    
    # Recent transactions
    recent_transactions = Transaction.objects.filter(
        team__in=user_teams
    ).order_by('-transaction_date')[:15]
    
    # Account summary by currency
    accounts_by_currency = {}
    for account in Account.objects.filter(is_active=True):
        currency = account.currency.code
        if currency not in accounts_by_currency:
            accounts_by_currency[currency] = {
                'total': 0,
                'accounts': []
            }
        accounts_by_currency[currency]['total'] += account.current_balance
        accounts_by_currency[currency]['accounts'].append(account)
    
    # Monthly summary (current month)
    current_month_start = datetime.now().replace(day=1)
    monthly_transactions = Transaction.objects.filter(
        team__in=user_teams,
        transaction_date__gte=current_month_start
    )
    
    monthly_income = monthly_transactions.filter(
        transaction_type='income'
    ).aggregate(total=Sum('amount_pkr'))['total'] or 0
    
    monthly_expense = monthly_transactions.filter(
        transaction_type='expense'
    ).aggregate(total=Sum('amount_pkr'))['total'] or 0
    
    monthly_net = monthly_income - monthly_expense
    
    # Team-wise expense summary
    team_expenses = []
    for team in user_teams:
        team_expense_total = Transaction.objects.filter(
            team=team,
            transaction_type='expense',
            transaction_date__gte=current_month_start
        ).aggregate(total=Sum('amount_pkr'))['total'] or 0
        
        # Get top expense category for this team
        top_category = Transaction.objects.filter(
            team=team,
            transaction_type='expense',
            transaction_date__gte=current_month_start
        ).values('category__name').annotate(
            total=Sum('amount_pkr')
        ).order_by('-total').first()
        
        team_expenses.append({
            'team_name': team.name,
            'total_expense': team_expense_total,
            'top_category': top_category['category__name'] if top_category else None
        })
    
    # Get all currencies for display
    currencies = Currency.objects.all().order_by('code')
    
    # Get USD currency for exchange rate display
    try:
        usd_currency = Currency.objects.get(code='USD')
    except Currency.DoesNotExist:
        usd_currency = None
    
    # Calculate totals for income/expense cards
    total_income = Transaction.objects.filter(
        transaction_type='income'
    ).aggregate(total=Sum('amount_pkr'))['total'] or 0
    
    total_expense = Transaction.objects.filter(
        transaction_type='expense'
    ).aggregate(total=Sum('amount_pkr'))['total'] or 0
    
    profit_loss = total_income - total_expense
    
    context = {
        'total_balance_pkr': total_balance_pkr,
        'recent_transactions': recent_transactions,
        'accounts_by_currency': accounts_by_currency,
        'user_teams': user_teams,
        'monthly_income': monthly_income,
        'monthly_expense': monthly_expense,
        'monthly_net': monthly_net,
        'team_expenses': team_expenses,
        'currencies': currencies,
        'usd_currency': usd_currency,
        'total_income': total_income,
        'total_expense': total_expense,
        'profit_loss': profit_loss,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def accounts_list(request):
    """List all accounts for user's teams"""
    user_teams = request.user.teams.all()
    accounts = Account.objects.filter(is_active=True).order_by('name')
    
    # Get currencies and account types for the modals
    currencies = Currency.objects.filter(is_active=True).order_by('code')
    account_types = Account.ACCOUNT_TYPES
    
    context = {
        'accounts': accounts,
        'user_teams': user_teams,
        'currencies': currencies,
        'account_types': account_types,
    }
    return render(request, 'core/accounts_list.html', context)


@login_required
def account_detail(request, account_id):
    """Account detail view with transactions"""
    account = get_object_or_404(Account, id=account_id)
    
    # Handle AJAX request for modal view - return just account details
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        account_data = {
            'name': account.name,
            'account_type': dict(Account.ACCOUNT_TYPES)[account.account_type],
            'currency': f"{account.currency.code} - {account.currency.name}",
            'opening_balance': f"{account.opening_balance:,.2f} {account.currency.code}",
            'current_balance': f"{account.current_balance:,.2f} {account.currency.code}",
            'opening_balance_pkr': f"{account.opening_balance_pkr:,.2f} PKR",
            'current_balance_pkr': f"{account.current_balance_pkr:,.2f} PKR",
            'description': account.description or 'No description provided',
            'is_active': 'Active' if account.is_active else 'Inactive',
            'created_at': account.created_at.strftime('%B %d, %Y at %I:%M %p'),
        }
        return JsonResponse(account_data)
    
    # For regular page requests, show full account detail with transactions
    transactions = Transaction.objects.filter(
        Q(account=account) | Q(counter_party_account=account)
    ).order_by('-transaction_date')
    
    context = {
        'account': account,
        'transactions': transactions,
    }
    return render(request, 'core/account_detail.html', context)


@login_required
def add_transaction(request):
    """Add new transaction"""
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                transaction = form.save(commit=False)
                transaction.created_by = request.user
                
                # Ensure amount_pkr is set if not already - use Decimal for calculations
                if not transaction.amount_pkr:
                    from decimal import Decimal
                    if transaction.currency and transaction.exchange_rate_to_pkr:
                        amount_decimal = Decimal(str(transaction.amount))
                        rate_decimal = Decimal(str(transaction.exchange_rate_to_pkr))
                        transaction.amount_pkr = amount_decimal * rate_decimal
                    else:
                        transaction.amount_pkr = Decimal(str(transaction.amount))  # Default to PKR
                
                transaction.save()
                messages.success(request, 'Transaction added successfully!')
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Transaction added successfully!'})
                return redirect('dashboard')
            except Exception as e:
                error_message = str(e)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': f'Error saving transaction: {error_message}'})
                messages.error(request, f'Error saving transaction: {error_message}')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = TransactionForm(user=request.user)
    
    # Helper function to get full category path (only for expense categories with parents)
    def get_category_display_name(cat):
        """Get full path of category for display"""
        if cat.category_type == 'expense' and cat.parent:
            parent_name = get_category_display_name(cat.parent) if cat.parent.parent else cat.parent.name
            return f"{parent_name} > {cat.name}"
        return cat.name
    
    # Get main categories and subcategories by type for JavaScript filtering
    main_categories_by_type = {
        'income': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='income', parent__isnull=True, is_active=True)],
        'expense': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='expense', parent__isnull=True, is_active=True)],
        'transfer': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='transfer', parent__isnull=True, is_active=True)],
        'owners_equity': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='owners_equity', parent__isnull=True, is_active=True)],
    }
    
    # Get subcategories with parent info for expense type
    subcategories_by_parent = {}
    for parent_cat in Category.objects.filter(category_type='expense', parent__isnull=True, is_active=True):
        subcategories_by_parent[parent_cat.id] = [
            {'id': sub.id, 'name': sub.name} 
            for sub in parent_cat.subcategories.filter(is_active=True)
        ]
    
    # Get all active accounts with currency info for JavaScript
    accounts = Account.objects.filter(is_active=True)
    accounts_data = [
        {
            'id': acc.id, 
            'name': acc.name, 
            'currency': acc.currency.code if acc.currency else 'PKR',
            'exchange_rate': float(acc.currency.exchange_rate_to_pkr) if acc.currency else 1.0000
        } 
        for acc in accounts
    ]
    
    # Get user's teams for JavaScript
    teams = Team.objects.all()
    teams_data = [
        {
            'id': team.id,
            'name': team.name
        }
        for team in teams
    ]
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form_html = render_to_string('core/transaction_form.html', {
            'form': form,
            'main_categories_by_type': json.dumps(main_categories_by_type),
            'subcategories_by_parent': json.dumps(subcategories_by_parent),
            'accounts_data': json.dumps(accounts_data),
            'teams_data': json.dumps(teams_data),
        }, request=request)
        return JsonResponse({'success': True, 'form_html': form_html})
    
    context = {
        'form': form,
        'main_categories_by_type': json.dumps(main_categories_by_type),
        'subcategories_by_parent': json.dumps(subcategories_by_parent),
        'accounts_data': json.dumps(accounts_data),
        'teams_data': json.dumps(teams_data),
    }
    return render(request, 'core/add_transaction.html', context)


@login_required
def transactions_list(request):
    """List all transactions"""
    # Show all transactions instead of filtering by user teams
    transactions = Transaction.objects.all().order_by('-transaction_date')
    
    # Filter by transaction type if specified
    transaction_type = request.GET.get('type')
    if transaction_type in ['income', 'expense', 'transfer']:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    # Filter by team if specified
    team_id = request.GET.get('team')
    if team_id:
        transactions = transactions.filter(team_id=team_id)
    
    # Get all teams for filter options
    all_teams = Team.objects.all()
    
    context = {
        'transactions': transactions,
        'user_teams': all_teams,  # Pass all teams instead of just user teams
        'current_type': transaction_type,
        'current_team': int(team_id) if team_id else None,
    }
    return render(request, 'core/transactions_list.html', context)


@login_required
def view_transaction(request, transaction_id):
    """View transaction details"""
    transaction = get_object_or_404(Transaction, id=transaction_id)
    
    # Allow viewing any transaction (removed team restriction)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'transaction': {
                'id': transaction.id,
                'transaction_type': transaction.get_transaction_type_display(),
                'description': transaction.description,
                'amount': str(transaction.amount),
                'amount_pkr': str(transaction.amount_pkr),
                'transaction_date': transaction.transaction_date.strftime('%B %d, %Y at %I:%M %p'),
                'account': transaction.account.name,
                'counter_party_account': transaction.counter_party_account.name if transaction.counter_party_account else None,
                'category': transaction.category.name if transaction.category else None,
                'team': transaction.team.name if transaction.team else None,
                'notes': transaction.notes,
                'currency': transaction.currency.code if transaction.currency else 'PKR',
                'exchange_rate_to_pkr': str(transaction.exchange_rate_to_pkr) if transaction.exchange_rate_to_pkr else None,
                'created_by': transaction.created_by.username,
                'created_at': transaction.created_at.strftime('%B %d, %Y at %I:%M %p'),
            }
        })
    
    context = {'transaction': transaction}
    return render(request, 'core/view_transaction.html', context)


@login_required
def edit_transaction(request, transaction_id):
    """Edit existing transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id)
    
    # Allow editing any transaction (removed team restriction)
    
    if request.method == 'POST':
        # Store original transaction data before editing
        original_data = {
            'transaction_type': transaction.transaction_type,
            'amount': transaction.amount,
            'amount_pkr': transaction.amount_pkr,
            'account': transaction.account,
            'counter_party_account': transaction.counter_party_account,
            'counter_party_amount': transaction.counter_party_amount,
            'counter_party_exchange_rate': transaction.counter_party_exchange_rate,
        }
        
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            try:
                # First, reverse the original transaction's effect on balances
                if original_data['transaction_type'] == 'income':
                    original_data['account'].current_balance -= original_data['amount']
                    original_data['account'].current_balance_pkr -= original_data['amount_pkr']
                    original_data['account'].save()
                    
                elif original_data['transaction_type'] == 'expense':
                    original_data['account'].current_balance += original_data['amount']
                    original_data['account'].current_balance_pkr += original_data['amount_pkr']
                    original_data['account'].save()
                    
                elif original_data['transaction_type'] == 'owners_equity':
                    # Owners equity was subtracted (like expense), so add it back
                    original_data['account'].current_balance += original_data['amount']
                    original_data['account'].current_balance_pkr += original_data['amount_pkr']
                    original_data['account'].save()
                    
                elif original_data['transaction_type'] == 'transfer':
                    if original_data['account'] and original_data['counter_party_account']:
                        # Reverse: add back to source account
                        original_data['account'].current_balance += original_data['amount']
                        original_data['account'].current_balance_pkr += original_data['amount_pkr']
                        
                        # Reverse: subtract from destination account
                        if original_data['counter_party_amount']:
                            original_data['counter_party_account'].current_balance -= original_data['counter_party_amount']
                            counter_party_pkr = original_data['counter_party_amount'] * original_data['counter_party_exchange_rate']
                            original_data['counter_party_account'].current_balance_pkr -= counter_party_pkr
                        else:
                            original_data['counter_party_account'].current_balance -= original_data['amount']
                            original_data['counter_party_account'].current_balance_pkr -= original_data['amount_pkr']
                        
                        original_data['account'].save()
                        original_data['counter_party_account'].save()
                
                # Save the updated transaction without triggering balance updates
                updated_transaction = form.save()
                
                # Manually apply the new transaction's balance changes
                updated_transaction.apply_balance_changes()
                
                messages.success(request, 'Transaction updated successfully and account balances recalculated!')
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Transaction updated successfully and account balances recalculated!'})
                return redirect('view_transaction', transaction_id=transaction.id)
                
            except Exception as e:
                error_message = f'Error updating transaction: {str(e)}'
                messages.error(request, error_message)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_message})
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    
    # Helper function to get full category path (only for expense categories with parents)
    def get_category_display_name(cat):
        """Get full path of category for display"""
        if cat.category_type == 'expense' and cat.parent:
            parent_name = get_category_display_name(cat.parent) if cat.parent.parent else cat.parent.name
            return f"{parent_name} > {cat.name}"
        return cat.name
    
    # Get main categories and subcategories by type for JavaScript filtering
    main_categories_by_type = {
        'income': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='income', parent__isnull=True, is_active=True)],
        'expense': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='expense', parent__isnull=True, is_active=True)],
        'transfer': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='transfer', parent__isnull=True, is_active=True)],
        'owners_equity': [{'id': cat.id, 'name': cat.name} for cat in Category.objects.filter(category_type='owners_equity', parent__isnull=True, is_active=True)],
    }
    
    # Get subcategories with parent info for expense type
    subcategories_by_parent = {}
    for parent_cat in Category.objects.filter(category_type='expense', parent__isnull=True, is_active=True):
        subcategories_by_parent[parent_cat.id] = [
            {'id': sub.id, 'name': sub.name} 
            for sub in parent_cat.subcategories.filter(is_active=True)
        ]
    
    # Get all active accounts with currency info for JavaScript
    accounts = Account.objects.filter(is_active=True)
    accounts_data = [
        {
            'id': acc.id, 
            'name': acc.name, 
            'currency': acc.currency.code if acc.currency else 'PKR',
            'exchange_rate': float(acc.currency.exchange_rate_to_pkr) if acc.currency else 1.0000
        } 
        for acc in accounts
    ]
    
    # Get all teams for JavaScript (not user-specific)
    teams = Team.objects.all()
    teams_data = [
        {
            'id': team.id,
            'name': team.name
        }
        for team in teams
    ]
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form_html = render_to_string('core/transaction_form.html', {
            'form': form,
            'transaction': transaction,
            'main_categories_by_type': json.dumps(main_categories_by_type),
            'subcategories_by_parent': json.dumps(subcategories_by_parent),
            'accounts_data': json.dumps(accounts_data),
            'teams_data': json.dumps(teams_data),
            'is_edit': True,
        }, request=request)
        return JsonResponse({'success': True, 'form_html': form_html})
    
    context = {
        'form': form,
        'transaction': transaction,
        'main_categories_by_type': json.dumps(main_categories_by_type),
        'subcategories_by_parent': json.dumps(subcategories_by_parent),
        'accounts_data': json.dumps(accounts_data),
        'teams_data': json.dumps(teams_data),
        'is_edit': True,
    }
    return render(request, 'core/edit_transaction.html', context)


@login_required
def delete_transaction(request, transaction_id):
    """Delete transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id)
    
    # Allow deleting any transaction (removed team restriction)
    
    if request.method == 'POST':
        # Reverse the transaction's effect on account balances before deleting
        try:
            if transaction.transaction_type == 'income':
                # Income was added to account, so subtract it back
                transaction.account.current_balance -= transaction.amount
                transaction.account.current_balance_pkr -= transaction.amount_pkr
                transaction.account.save()
                
            elif transaction.transaction_type == 'expense':
                # Expense was subtracted from account, so add it back
                transaction.account.current_balance += transaction.amount
                transaction.account.current_balance_pkr += transaction.amount_pkr
                transaction.account.save()
                
            elif transaction.transaction_type == 'owners_equity':
                # Owners equity was subtracted from account (like expense), so add it back
                transaction.account.current_balance += transaction.amount
                transaction.account.current_balance_pkr += transaction.amount_pkr
                transaction.account.save()
                
            elif transaction.transaction_type == 'transfer':
                # Transfer moved money between accounts, reverse both
                if transaction.account and transaction.counter_party_account:
                    # Reverse: add back to source account (it was debited)
                    transaction.account.current_balance += transaction.amount
                    transaction.account.current_balance_pkr += transaction.amount_pkr
                    
                    # Reverse: subtract from destination account (it was credited)
                    if transaction.counter_party_amount:
                        transaction.counter_party_account.current_balance -= transaction.counter_party_amount
                        counter_party_pkr = transaction.counter_party_amount * transaction.counter_party_exchange_rate
                        transaction.counter_party_account.current_balance_pkr -= counter_party_pkr
                    else:
                        # Fallback: use same currency amounts
                        transaction.counter_party_account.current_balance -= transaction.amount
                        transaction.counter_party_account.current_balance_pkr -= transaction.amount_pkr
                    
                    transaction.account.save()
                    transaction.counter_party_account.save()
            
            # Delete the transaction
            transaction.delete()
            
            messages.success(request, 'Transaction deleted successfully and account balances updated!')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Transaction deleted successfully and account balances updated!'})
                
        except Exception as e:
            error_message = f'Error deleting transaction: {str(e)}'
            messages.error(request, error_message)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_message})
                
        return redirect('transactions_list')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'transaction': {
                'id': transaction.id,
                'description': transaction.description,
                'amount': str(transaction.amount),
                'transaction_type': transaction.get_transaction_type_display(),
                'account': transaction.account.name,
            }
        })
    
    context = {'transaction': transaction}
    return render(request, 'core/delete_transaction.html', context)


@login_required
def categories_list(request):
    """List all categories organized hierarchically"""
    user_teams = request.user.teams.all()
    
    # Get all categories organized by type
    categories_by_type = {
        'income': [],
        'expense': [],
        'transfer': [],
        'owners_equity': []
    }
    
    # Get main categories (those without parents)
    main_categories = Category.objects.filter(is_active=True, parent=None).order_by('category_type', 'name')
    
    # For each main category, get its hierarchy
    for main_cat in main_categories:
        cat_type = main_cat.category_type
        if cat_type in categories_by_type:
            # Build hierarchy
            hierarchy = {
                'category': main_cat,
                'children': get_category_hierarchy(main_cat)
            }
            categories_by_type[cat_type].append(hierarchy)
    
    # Get only main categories (no subcategories) for dropdown filter
    main_categories_list = Category.objects.filter(is_active=True, parent__isnull=True).order_by('category_type', 'name')
    
    context = {
        'categories_by_type': categories_by_type,
        'all_categories': main_categories_list,
        'user_teams': user_teams,
    }
    return render(request, 'core/categories_list.html', context)


def get_category_hierarchy(category):
    """Recursively build category hierarchy"""
    children = []
    for child in category.subcategories.filter(is_active=True).order_by('name'):
        children.append({
            'category': child,
            'children': get_category_hierarchy(child)
        })
    return children


@login_required
def add_account(request):
    """Add new account"""
    if request.method == 'POST':
        form = AccountForm(request.POST, user=request.user)
        if form.is_valid():
            account = form.save()
            
            # Handle AJAX request
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Account added successfully!'})
            
            messages.success(request, 'Account added successfully!')
            return redirect('accounts_list')
        else:
            # Handle AJAX form errors
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = AccountForm(user=request.user)
    
    context = {'form': form}
    return render(request, 'core/add_account.html', context)


@login_required
def edit_account(request, account_id):
    """Edit existing account"""
    account = get_object_or_404(Account, id=account_id)
    
    # Handle AJAX request for getting account data
    if request.method == 'GET' and request.META.get('HTTP_ACCEPT') == 'application/json':
        return JsonResponse({
            'id': account.id,
            'name': account.name,
            'account_type': account.account_type,
            'currency': account.currency.id if account.currency else '',
            'opening_balance': str(account.opening_balance),
            'description': account.description or '',
            'is_active': account.is_active,
        })
    
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account, user=request.user)
        if form.is_valid():
            account = form.save()
            
            # Handle AJAX request
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Account updated successfully!'})
            
            messages.success(request, 'Account updated successfully!')
            return redirect('accounts_list')
        else:
            # Handle AJAX form errors
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = AccountForm(instance=account, user=request.user)
    
    context = {
        'form': form,
        'account': account,
        'is_edit': True
    }
    return render(request, 'core/add_account.html', context)


@login_required
def delete_account(request, account_id):
    """Delete account"""
    account = get_object_or_404(Account, id=account_id)
    
    # Check if account has transactions
    has_transactions = Transaction.objects.filter(
        Q(account=account) | Q(counter_party_account=account)
    ).exists()
    
    if has_transactions:
        # Handle AJAX request
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete account "{account.name}" because it has transactions.'
            }, status=400)
        
        messages.error(request, f'Cannot delete account "{account.name}" because it has transactions.')
        return redirect('accounts_list')
    
    if request.method == 'POST':
        account_name = account.name
        account.delete()
        
        # Handle AJAX request
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Account "{account_name}" deleted successfully!'})
        
        messages.success(request, f'Account "{account_name}" deleted successfully!')
        return redirect('accounts_list')
    
    context = {'account': account}
    return render(request, 'core/delete_account.html', context)


@login_required
def add_category(request):
    """Add new category"""
    if request.method == 'POST':
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Category added successfully!')
            return redirect('categories_list')
    else:
        form = CategoryForm(user=request.user)
    
    context = {'form': form}
    return render(request, 'core/add_category.html', context)


@login_required
def edit_category(request, category_id):
    """Edit existing category"""
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category, user=request.user)
        if form.is_valid():
            category = form.save()
            messages.success(request, 'Category updated successfully!')
            return redirect('categories_list')
    else:
        form = CategoryForm(instance=category, user=request.user)
    
    context = {
        'form': form,
        'category': category,
        'is_edit': True
    }
    return render(request, 'core/add_category.html', context)


@login_required
def delete_category(request, category_id):
    """Delete category"""
    category = get_object_or_404(Category, id=category_id)
    
    # Check if category has subcategories
    if category.subcategories.filter(is_active=True).exists():
        messages.error(request, f'Cannot delete category "{category.name}" because it has subcategories. Delete subcategories first.')
        return redirect('categories_list')
    
    # Check if category is being used in transactions
    if category.transaction_set.exists():
        messages.error(request, f'Cannot delete category "{category.name}" because it is being used in transactions.')
        return redirect('categories_list')
    
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Category "{category_name}" deleted successfully!')
        return redirect('categories_list')
    
    context = {'category': category}
    return render(request, 'core/delete_category.html', context)


@login_required
def teams_list(request):
    """List all teams"""
    teams = Team.objects.all().order_by('name')
    
    context = {
        'teams': teams,
    }
    return render(request, 'core/teams_list.html', context)


@login_required
def team_detail(request, team_id):
    """Team detail view"""
    team = get_object_or_404(Team, id=team_id)
    
    # Handle AJAX request for modal view - return just team details
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        team_data = {
            'name': team.name,
            'description': team.description or 'No description provided',
            'created_at': team.created_at.strftime('%B %d, %Y at %I:%M %p'),
        }
        return JsonResponse(team_data)
    
    # For regular page requests, show full team detail
    context = {
        'team': team,
    }
    return render(request, 'core/team_detail.html', context)


@login_required
def add_team(request):
    """Add new team"""
    if request.method == 'POST':
        form = TeamForm(request.POST, user=request.user)
        if form.is_valid():
            team = form.save(commit=False)
            team.created_by = request.user
            team.save()
            
            # Handle AJAX request
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Team added successfully!'})
            
            messages.success(request, 'Team added successfully!')
            return redirect('teams_list')
        else:
            # Handle AJAX form errors
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = TeamForm(user=request.user)
    
    context = {'form': form}
    return render(request, 'core/add_team.html', context)


@login_required
def edit_team(request, team_id):
    """Edit existing team"""
    team = get_object_or_404(Team, id=team_id)
    
    # Handle AJAX request for getting team data
    if request.method == 'GET' and request.META.get('HTTP_ACCEPT') == 'application/json':
        return JsonResponse({
            'id': team.id,
            'name': team.name,
            'description': team.description or '',
        })
    
    if request.method == 'POST':
        form = TeamForm(request.POST, instance=team, user=request.user)
        if form.is_valid():
            team = form.save()
            
            # Handle AJAX request
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Team updated successfully!'})
            
            messages.success(request, 'Team updated successfully!')
            return redirect('teams_list')
        else:
            # Handle AJAX form errors
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = TeamForm(instance=team, user=request.user)
    
    context = {
        'form': form,
        'team': team,
        'is_edit': True
    }
    return render(request, 'core/add_team.html', context)


@login_required
def delete_team(request, team_id):
    """Delete team"""
    team = get_object_or_404(Team, id=team_id)
    
    # Check if team has transactions
    has_transactions = Transaction.objects.filter(team=team).exists()
    
    if has_transactions:
        # Handle AJAX request
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete team "{team.name}" because it has transactions.'
            }, status=400)
        
        messages.error(request, f'Cannot delete team "{team.name}" because it has transactions.')
        return redirect('teams_list')
    
    if request.method == 'POST':
        team_name = team.name
        team.delete()
        
        # Handle AJAX request
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Team "{team_name}" deleted successfully!'})
        
        messages.success(request, f'Team "{team_name}" deleted successfully!')
        return redirect('teams_list')
    
    context = {'team': team}
    return render(request, 'core/delete_team.html', context)


@login_required
def reports(request):
    """Reports view showing financial analytics and reports with filtering and download"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    import csv
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Helper function to generate CSV report
    def generate_csv_report(transactions, filter_params):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Date', 'Transaction Type', 'Category', 'Subcategory', 'Account', 'Counter Party', 'Team', 
            'Currency', 'Exchange Rate', 'Original Amount', 'PKR Amount', 'Description'
        ])
        
        # Write data
        for transaction in transactions:
            # Determine category and subcategory
            category_name = '-'
            subcategory_name = '-'
            if transaction.category:
                if transaction.category.parent:
                    category_name = transaction.category.parent.name
                    subcategory_name = transaction.category.name
                else:
                    category_name = transaction.category.name
            
            # Format exchange rate
            exchange_rate = f"{transaction.exchange_rate_to_pkr:.4f}" if transaction.exchange_rate_to_pkr else "1.0000"
            
            writer.writerow([
                transaction.transaction_date.strftime('%Y-%m-%d'),
                transaction.transaction_type.title(),
                category_name,
                subcategory_name,
                transaction.account.name,
                transaction.counter_party_account.name if transaction.counter_party_account else '-',
                transaction.team.name if transaction.team else '-',
                transaction.currency.code,
                exchange_rate,
                f"{transaction.amount:.2f}",
                f"{transaction.amount_pkr:.2f}",
                transaction.description
            ])
        
        return response
    
    # Helper function to generate Excel report
    def generate_excel_report(transactions, filter_params):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Date', 'Transaction Type', 'Category', 'Subcategory', 'Account', 'Counter Party', 'Team', 
            'Currency', 'Exchange Rate', 'Original Amount', 'PKR Amount', 'Description'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, transaction in enumerate(transactions, 2):
            # Determine category and subcategory
            category_name = '-'
            subcategory_name = '-'
            if transaction.category:
                if transaction.category.parent:
                    category_name = transaction.category.parent.name
                    subcategory_name = transaction.category.name
                else:
                    category_name = transaction.category.name
            
            ws.cell(row=row, column=1, value=transaction.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=transaction.transaction_type.title())
            ws.cell(row=row, column=3, value=category_name)
            ws.cell(row=row, column=4, value=subcategory_name)
            ws.cell(row=row, column=5, value=transaction.account.name)
            ws.cell(row=row, column=6, value=transaction.counter_party_account.name if transaction.counter_party_account else '-')
            ws.cell(row=row, column=7, value=transaction.team.name if transaction.team else '-')
            ws.cell(row=row, column=8, value=transaction.currency.code)
            ws.cell(row=row, column=9, value=float(transaction.exchange_rate_to_pkr) if transaction.exchange_rate_to_pkr else 1.0000)
            ws.cell(row=row, column=10, value=float(transaction.amount))
            ws.cell(row=row, column=11, value=float(transaction.amount_pkr))
            ws.cell(row=row, column=12, value=transaction.description)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    # Check for download request
    if request.GET.get('download') == 'true':
        format_type = request.GET.get('format', 'csv')
        
        # Get filtered transactions for download
        transactions = Transaction.objects.all().order_by('-transaction_date')
        
        # Apply filters for download
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        transaction_type = request.GET.get('transaction_type')
        category_id = request.GET.get('category')
        subcategory_id = request.GET.get('subcategory')
        account_id = request.GET.get('account')
        team_id = request.GET.get('team')
        
        if start_date:
            transactions = transactions.filter(transaction_date__gte=start_date)
        if end_date:
            transactions = transactions.filter(transaction_date__lte=end_date)
        if transaction_type and transaction_type != 'all':
            transactions = transactions.filter(transaction_type=transaction_type)
        if category_id and category_id != 'all':
            transactions = transactions.filter(
                Q(category_id=category_id) | Q(category__parent_id=category_id)
            )
        if subcategory_id and subcategory_id != 'all':
            transactions = transactions.filter(category_id=subcategory_id)
        if account_id and account_id != 'all':
            transactions = transactions.filter(account_id=account_id)
        if team_id and team_id != 'all':
            transactions = transactions.filter(team_id=team_id)
        
        filter_params = {
            'start_date': start_date,
            'end_date': end_date,
            'transaction_type': transaction_type,
            'category': category_id,
            'account': account_id,
            'team': team_id,
        }
        
        if format_type == 'excel':
            return generate_excel_report(transactions, filter_params)
        else:
            return generate_csv_report(transactions, filter_params)
    
    # Get filter parameters
    current_start_date = request.GET.get('start_date', '')
    current_end_date = request.GET.get('end_date', '')
    current_transaction_type = request.GET.get('transaction_type', 'all')
    current_category = request.GET.get('category', 'all')
    current_subcategory = request.GET.get('subcategory', 'all')
    current_account = request.GET.get('account', 'all')
    current_team = request.GET.get('team', 'all')
    
    # Start with all transactions
    transactions = Transaction.objects.all()
    
    # Apply filters
    if current_start_date:
        transactions = transactions.filter(transaction_date__gte=current_start_date)
    if current_end_date:
        transactions = transactions.filter(transaction_date__lte=current_end_date)
    if current_transaction_type != 'all':
        transactions = transactions.filter(transaction_type=current_transaction_type)
    if current_category != 'all':
        try:
            current_category = int(current_category)
            # Filter by main category - get both the main category and its subcategories
            transactions = transactions.filter(
                Q(category_id=current_category) | Q(category__parent_id=current_category)
            )
        except (ValueError, TypeError):
            current_category = 'all'
    if current_subcategory != 'all':
        try:
            current_subcategory = int(current_subcategory)
            transactions = transactions.filter(category_id=current_subcategory)
        except (ValueError, TypeError):
            current_subcategory = 'all'
    if current_account != 'all':
        try:
            current_account = int(current_account)
            transactions = transactions.filter(account_id=current_account)
        except (ValueError, TypeError):
            current_account = 'all'
    if current_team != 'all':
        try:
            current_team = int(current_team)
            transactions = transactions.filter(team_id=current_team)
        except (ValueError, TypeError):
            current_team = 'all'
    
    # Get consolidated breakdown (individual transactions)
    consolidated_breakdown = transactions.order_by('-transaction_date')
    
    # Calculate totals
    total_income = transactions.filter(transaction_type='income').aggregate(
        total=Sum('amount_pkr'))['total'] or Decimal('0')
    total_expense = transactions.filter(transaction_type='expense').aggregate(
        total=Sum('amount_pkr'))['total'] or Decimal('0')
    total_transfers = transactions.filter(transaction_type='transfer').aggregate(
        total=Sum('amount_pkr'))['total'] or Decimal('0')
    
    net_amount = total_income - total_expense
    
    # Get all filter options
    all_categories = Category.objects.filter(is_active=True, parent__isnull=True).order_by('category_type', 'name')
    all_subcategories = Category.objects.filter(is_active=True, parent__isnull=False).order_by('parent__name', 'name')
    all_accounts = Account.objects.filter(is_active=True).order_by('name')
    all_teams = Team.objects.all().order_by('name')
    
    context = {
        'consolidated_breakdown': consolidated_breakdown,
        'total_income': total_income,
        'total_expense': total_expense,
        'total_transfers': total_transfers,
        'net_amount': net_amount,
        'all_categories': all_categories,
        'all_subcategories': all_subcategories,
        'all_accounts': all_accounts,
        'all_teams': all_teams,
        'current_start_date': current_start_date,
        'current_end_date': current_end_date,
        'current_transaction_type': current_transaction_type,
        'current_category': current_category,
        'current_subcategory': current_subcategory,
        'current_account': current_account,
        'current_team': current_team,
    }
    
    return render(request, 'core/reports.html', context)


@login_required
def update_exchange_rates(request):
    """Update exchange rates and recalculate account balances"""
    if request.method == 'POST':
        # Handle form submission to update rates
        updated_currencies = []
        
        for currency in Currency.objects.exclude(code='PKR'):
            rate_key = f'rate_{currency.code}'
            if rate_key in request.POST:
                try:
                    new_rate = Decimal(request.POST[rate_key])
                    if new_rate > 0:
                        old_rate = currency.exchange_rate_to_pkr
                        currency.exchange_rate_to_pkr = new_rate
                        currency.save()
                        updated_currencies.append(currency.code)
                        
                        # Recalculate balances for all accounts with this currency
                        accounts = Account.objects.filter(currency=currency)
                        for account in accounts:
                            # Recalculate the current balance from scratch
                            account.calculate_current_balance()
                            account.save()
                        
                        # Update all transaction PKR amounts for this currency
                        transactions = Transaction.objects.filter(currency=currency)
                        for transaction in transactions:
                            transaction.amount_pkr = transaction.amount * new_rate
                            transaction.save(update_fields=['amount_pkr'])
                        
                        # Update counter party transactions (for transfers)
                        counter_transactions = Transaction.objects.filter(counter_party_currency=currency)
                        for transaction in counter_transactions:
                            # Recalculate counter party amount with new rate
                            if transaction.currency and transaction.exchange_rate_to_pkr:
                                amount_pkr = transaction.amount * transaction.exchange_rate_to_pkr
                                transaction.counter_party_amount = amount_pkr / new_rate
                                transaction.save(update_fields=['counter_party_amount'])
                        
                        messages.success(request, f'Updated {currency.code} rate from {old_rate} to {new_rate} PKR and recalculated all balances')
                    else:
                        messages.error(request, f'Invalid rate for {currency.code}')
                except (ValueError, TypeError):
                    messages.error(request, f'Invalid rate format for {currency.code}')
        
        if updated_currencies:
            messages.success(request, f'Exchange rates updated successfully! Recalculated balances for: {", ".join(updated_currencies)}')
        
        return redirect('accounts_list')
    
    # GET request - redirect to accounts page
    return redirect('accounts_list')
